"""
ERP Data Cleaning & Anomaly Detection Pipeline
Entry point – reads online_retail.xlsx, cleans, detects anomalies, exports Excel report.
"""
import os
import sys
import warnings
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from sklearn.ensemble import IsolationForest

warnings.filterwarnings("ignore")

# ── CONFIG FROM ENV ──────────────────────────────────────────────────────────
DATA_DIR   = Path(os.getenv("DATA_DIR",   "/app/data"))
OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", "/app/output"))
LOG_LEVEL  = os.getenv("LOG_LEVEL", "INFO")

IQR_MULT       = float(os.getenv("IQR_MULTIPLIER",      "1.5"))
ZSCORE_THR     = float(os.getenv("ZSCORE_THRESHOLD",    "3.0"))
ISO_CONTAM     = float(os.getenv("ISO_CONTAMINATION",   "0.03"))
ANOM_LIMIT     = int(os.getenv("ANOMALY_EXPORT_LIMIT",  "10000"))

INPUT_FILE = DATA_DIR / "online_retail.xlsx"

# ── LOGGER ───────────────────────────────────────────────────────────────────
logger.remove()
logger.add(
    sys.stdout,
    format="<green>{time:YYYY-MM-DD HH:mm:ss}</green> | <level>{level:<8}</level> | {message}",
    level=LOG_LEVEL,
    colorize=True,
)
logger.add(
    OUTPUT_DIR / "pipeline.log",
    format="{time:YYYY-MM-DD HH:mm:ss} | {level:<8} | {message}",
    level="DEBUG",
    rotation="5 MB",
)

# ── STYLE HELPERS ────────────────────────────────────────────────────────────
NAVY = "1F3864"; WHT = "FFFFFF"; BLU = "2E75B6"; ORG = "ED7D31"
RED  = "C00000"; GRN = "375623"; GRY = "F2F2F2"; PUR = "7030A0"
_thin = Side(style="thin", color="BFBFBF")
_bdr  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _hdr(ws, row: int, ncols: int, bg: str = NAVY, start: int = 1):
    for c in range(start, start + ncols):
        cell = ws.cell(row, c)
        cell.font      = Font(bold=True, color=WHT, size=10, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _bdr
    ws.row_dimensions[row].height = 28


def _drow(ws, row: int, ncols: int, alt: bool = False, start: int = 1):
    bg = GRY if alt else WHT
    for c in range(start, start + ncols):
        ws.cell(row, c).fill      = PatternFill("solid", fgColor=bg)
        ws.cell(row, c).border    = _bdr
        ws.cell(row, c).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, c).font      = Font(name="Calibri", size=10)


def _autofit(ws, mn: int = 10, mx: int = 38):
    for col in ws.columns:
        L = get_column_letter(col[0].column)
        w = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[L].width = min(max(w + 2, mn), mx)


def _write_df(ws, df_in: pd.DataFrame, sr: int = 1, bg: str = NAVY):
    cols = df_in.columns.tolist()
    for i, c in enumerate(cols, 1):
        ws.cell(sr, i, c)
    _hdr(ws, sr, len(cols), bg=bg)
    for ri, row_d in enumerate(df_in.itertuples(index=False), sr + 1):
        _drow(ws, ri, len(cols), alt=(ri % 2 == 0))
        for ci, val in enumerate(row_d, 1):
            cell = ws.cell(ri, ci, val)
            if isinstance(val, float):
                cell.number_format = "#,##0.00"
            elif isinstance(val, int):
                cell.number_format = "#,##0"


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1 – LOAD
# ═══════════════════════════════════════════════════════════════════════════════
def load_data(path: Path) -> pd.DataFrame:
    logger.info(f"Loading raw data from  →  {path}")
    if not path.exists():
        logger.error(f"Input file not found: {path}")
        sys.exit(1)
    df = pd.read_excel(path, engine="openpyxl")
    logger.info(f"Rows loaded            →  {len(df):,}")
    return df


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2 – CLEAN
# ═══════════════════════════════════════════════════════════════════════════════
def clean_data(df: pd.DataFrame) -> tuple[pd.DataFrame, list]:
    logger.info("── DATA CLEANING ──────────────────────────────────────")
    log = []  # (step_label, n_affected, pct, action)
    initial = len(df)

    # Standardise column names
    df.columns = [c.strip() for c in df.columns]
    logger.info(f"  [rename]     Columns: {df.columns.tolist()}")

    # Remove exact duplicates
    before = len(df)
    df = df.drop_duplicates()
    n = before - len(df)
    log.append(("Remove Exact Duplicates", n, f"{n/before*100:.2f}%", "Dropped"))
    logger.info(f"  [dedup]      Removed {n:,} exact duplicate rows")

    # Tag missing CustomerID as GUEST (keep row)
    n = int(df["CustomerID"].isna().sum())
    df["CustomerID"] = df["CustomerID"].fillna(0).astype(int)
    df["CustomerID"] = df["CustomerID"].astype(str).replace("0", "GUEST")
    log.append(("Tag Missing CustomerID as GUEST", n, f"{n/len(df)*100:.2f}%", "Tagged"))
    logger.info(f"  [customer]   Tagged {n:,} rows as GUEST (missing CustomerID)")

    # Remove cancellations (InvoiceNo starts with C)
    before = len(df)
    cancel_mask = df["InvoiceNo"].astype(str).str.startswith("C")
    df_cancelled = df[cancel_mask].copy()
    df = df[~cancel_mask]
    n = before - len(df)
    log.append(("Remove Cancellation Transactions", n, f"{n/before*100:.2f}%", "Dropped"))
    logger.info(f"  [cancel]     Removed {n:,} cancellation transactions (InvoiceNo starts with C)")

    # Correct negative quantity → abs (returns data, keep for analysis)
    neg_mask = df["Quantity"] < 0
    n = int(neg_mask.sum())
    df.loc[neg_mask, "Quantity"] = df.loc[neg_mask, "Quantity"].abs()
    log.append(("Correct Negative Quantity (abs)", n, f"{n/len(df)*100:.2f}%", "Corrected"))
    logger.info(f"  [quantity]   Corrected {n:,} negative-quantity rows (abs applied)")

    # Flag rows with suspiciously low unit price
    low_price_mask = df["UnitPrice"] < 0.01
    n = int(low_price_mask.sum())
    df["LowPrice_Flag"] = low_price_mask
    log.append(("Flag UnitPrice < 0.01", n, f"{n/len(df)*100:.2f}%", "Flagged"))
    logger.info(f"  [price]      Flagged {n:,} rows with UnitPrice < 0.01")

    # Parse InvoiceDate
    df["InvoiceDate"] = pd.to_datetime(df["InvoiceDate"], errors="coerce")
    bad_dates = int(df["InvoiceDate"].isna().sum())
    log.append(("Parse InvoiceDate", bad_dates, f"{bad_dates/len(df)*100:.2f}%", "Parsed"))
    logger.info(f"  [datetime]   Parsed InvoiceDate; {bad_dates} invalid dates found")

    # Strip whitespace from text columns
    text_cols = df.select_dtypes("object").columns.tolist()
    for col in text_cols:
        df[col] = df[col].astype(str).str.strip()
    logger.info(f"  [whitespace] Stripped whitespace from {len(text_cols)} text columns")

    # Add derived columns
    df["Revenue"]      = (df["Quantity"] * df["UnitPrice"]).round(2)
    df["Quarter"]      = df["InvoiceDate"].dt.to_period("Q").astype(str)
    df["InvoiceMonth"] = df["InvoiceDate"].dt.to_period("M").astype(str)
    df["DayOfWeek"]    = df["InvoiceDate"].dt.day_name()
    df["Hour"]         = df["InvoiceDate"].dt.hour
    log.append(("Add Derived Features", 0, "—", "Added Revenue, Quarter, Month, DayOfWeek, Hour"))
    logger.info("  [features]   Added Revenue, Quarter columns")

    final = len(df)
    logger.info(f"  Cleaning complete.  Final rows: {final:,}")
    return df, log


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 3 – FEATURE ENGINEERING
# ═══════════════════════════════════════════════════════════════════════════════
def engineer_features(df: pd.DataFrame) -> pd.DataFrame:
    logger.info("Engineering features (TotalPrice, OrderSize, CustomerRFM) …")

    order_size = df.groupby("InvoiceNo")["Quantity"].sum().rename("OrderSize")
    df = df.merge(order_size, on="InvoiceNo", how="left")

    # Simple RFM proxy per CustomerID
    ref_date = df["InvoiceDate"].max()
    rfm = df.groupby("CustomerID").agg(
        Recency=("InvoiceDate",  lambda x: (ref_date - x.max()).days),
        Frequency=("InvoiceNo", "nunique"),
        Monetary=("Revenue",    "sum"),
    ).reset_index()
    df = df.merge(rfm, on="CustomerID", how="left")
    return df


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 4 – ANOMALY DETECTION
# ═══════════════════════════════════════════════════════════════════════════════
def detect_anomalies(df: pd.DataFrame) -> tuple[pd.DataFrame, list]:
    logger.info("── ANOMALY DETECTION ───────────────────────────────────")
    anom_log = []

    def iqr_flag(col: str) -> pd.Series:
        s = df[col]
        Q1, Q3 = s.quantile(0.25), s.quantile(0.75)
        IQR = Q3 - Q1
        lo, hi = Q1 - IQR_MULT * IQR, Q3 + IQR_MULT * IQR
        mask = (s < lo) | (s > hi)
        n = int(mask.sum())
        logger.info(f"  [IQR] {col:<16} fence=[{lo:.2f}, {hi:.2f}]  flagged={n:,}")
        anom_log.append((f"IQR – {col}", lo, hi, n, f"{n/len(df)*100:.2f}%"))
        return mask

    def z_flag(col: str) -> pd.Series:
        s = df[col]
        z = (s - s.mean()) / s.std()
        mask = z.abs() > ZSCORE_THR
        n = int(mask.sum())
        logger.info(f"  [Z-Score] {col:<14} |z|>{ZSCORE_THR}  flagged={n:,}")
        anom_log.append((f"Z-Score – {col}", -ZSCORE_THR, ZSCORE_THR, n, f"{n/len(df)*100:.2f}%"))
        return mask

    df["A_IQR_Qty"]    = iqr_flag("Quantity")
    df["A_IQR_Price"]  = iqr_flag("UnitPrice")
    df["A_IQR_Rev"]    = iqr_flag("Revenue")
    df["A_Z_Qty"]      = z_flag("Quantity")
    df["A_Z_Price"]    = z_flag("UnitPrice")
    df["A_Z_Rev"]      = z_flag("Revenue")

    # Isolation Forest on numeric features
    features = ["Quantity", "UnitPrice", "Revenue", "OrderSize"]
    feat_df  = df[features].fillna(0)
    iso = IsolationForest(contamination=ISO_CONTAM, random_state=42, n_jobs=-1)
    df["A_IsoForest"] = iso.fit_predict(feat_df) == -1
    n_iso = int(df["A_IsoForest"].sum())
    logger.info(f"  [IsoForest] contamination={ISO_CONTAM}  flagged={n_iso:,}")
    anom_log.append((f"IsoForest (contamination={ISO_CONTAM})", "—", "—", n_iso, f"{n_iso/len(df)*100:.2f}%"))

    anom_cols = ["A_IQR_Qty","A_IQR_Price","A_IQR_Rev","A_Z_Qty","A_Z_Price","A_Z_Rev","A_IsoForest"]
    df["Is_Anomaly"]    = df[anom_cols].any(axis=1)
    df["Anomaly_Score"] = df[anom_cols].sum(axis=1).astype(int)

    total = int(df["Is_Anomaly"].sum())
    pct   = total / len(df) * 100
    logger.info(f"  Total anomalies  →  {total:,}  ({pct:.1f}%)")
    return df, anom_log, total, pct


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 5 – EXPORT
# ═══════════════════════════════════════════════════════════════════════════════
def export_outputs(df: pd.DataFrame, clean_log: list, anom_log: list,
                   total_anom: int, anom_pct: float, initial_rows: int):

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    df_clean = df[~df["Is_Anomaly"]]
    df_anom  = df[df["Is_Anomaly"]]
    clean_rows = len(df)

    # ── CSVs ─────────────────────────────────────────────────────────────────
    clean_csv = OUTPUT_DIR / f"erp_cleaned_{ts}.csv"
    anom_csv  = OUTPUT_DIR / f"anomalies_only_{ts}.csv"
    df_clean.to_csv(clean_csv, index=False)
    logger.info(f"  [export] CSV  →  {clean_csv}")
    df_anom.to_csv(anom_csv, index=False)
    logger.info(f"  [export] Anomalies CSV  →  {anom_csv}  ({len(df_anom):,} rows)")

    # ── EXCEL REPORT ──────────────────────────────────────────────────────────
    xlsx_path = OUTPUT_DIR / f"ERP_Anomaly_Report_{ts}.xlsx"

    # KPI tables
    monthly = df_clean.groupby("InvoiceMonth").agg(
        Total_Revenue=("Revenue","sum"), Invoices=("InvoiceNo","nunique"),
        Customers=("CustomerID","nunique"), Avg_Order=("Revenue","mean"),
        Items_Sold=("Quantity","sum"),
    ).reset_index().round(2)

    country = df_clean.groupby("Country").agg(
        Total_Revenue=("Revenue","sum"), Transactions=("InvoiceNo","count"),
        Avg_Price=("UnitPrice","mean"),
    ).reset_index().sort_values("Total_Revenue", ascending=False).round(2)

    top_products = df_clean.groupby(["StockCode","Description"]).agg(
        Qty_Sold=("Quantity","sum"), Revenue=("Revenue","sum"),
        Transactions=("InvoiceNo","count"), Avg_Price=("UnitPrice","mean"),
    ).reset_index().sort_values("Revenue", ascending=False).head(50).round(2)

    anom_month = df.groupby("InvoiceMonth")["Is_Anomaly"].agg(["sum","count"]).reset_index()
    anom_month.columns = ["Month","Anomaly_Count","Total_Count"]
    anom_month["Anomaly_Rate_%"] = (anom_month["Anomaly_Count"]/anom_month["Total_Count"]*100).round(2)

    # Anomaly export (capped)
    anom_exp = df_anom[[
        "InvoiceNo","StockCode","Description","Quantity","UnitPrice","Revenue",
        "InvoiceDate","Country","CustomerID",
        "A_IQR_Qty","A_IQR_Price","A_IQR_Rev","A_Z_Qty","A_Z_Price","A_Z_Rev",
        "A_IsoForest","Anomaly_Score",
    ]].head(ANOM_LIMIT).copy()
    anom_exp["InvoiceDate"] = anom_exp["InvoiceDate"].astype(str)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Executive Summary ────────────────────────────────────────────
    ws1 = wb.create_sheet("Executive_Summary")
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = BLU

    # Title
    ws1.merge_cells("A1:I1")
    ws1["A1"] = "ERP DATA CLEANING & ANOMALY DETECTION — EXECUTIVE SUMMARY"
    ws1["A1"].font      = Font(bold=True, size=15, color=WHT, name="Calibri")
    ws1["A1"].fill      = PatternFill("solid", fgColor=NAVY)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 44

    ws1.merge_cells("A2:I2")
    ws1["A2"] = (
        f"Pipeline Run: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  "
        f"Source: online_retail.xlsx  |  "
        f"IQR Multiplier: {IQR_MULT}  |  Z-Score Threshold: {ZSCORE_THR}  |  "
        f"IsoForest Contamination: {ISO_CONTAM}"
    )
    ws1["A2"].font      = Font(italic=True, size=9, color="595959", name="Calibri")
    ws1["A2"].fill      = PatternFill("solid", fgColor="D9E1F2")
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 18

    # KPI section
    ws1.merge_cells("A3:I3")
    ws1["A3"] = "  KEY PERFORMANCE INDICATORS"
    ws1["A3"].font = Font(bold=True, size=11, color=NAVY, name="Calibri")
    ws1["A3"].fill = PatternFill("solid", fgColor="D9E1F2")

    kpis = [
        ("RAW RECORDS",   f"{initial_rows:,}",    NAVY, ["A","B"]),
        ("AFTER CLEAN",   f"{clean_rows:,}",       GRN,  ["C","D"]),
        ("ANOMALIES",     f"{total_anom:,}",       RED,  ["E","F"]),
        ("ANOMALY RATE",  f"{anom_pct:.2f}%",      RED,  ["G","H"]),
        ("COUNTRIES",     f"{df_clean['Country'].nunique():,}", BLU, ["I","I"]),
    ]
    for label, val, color, cols in kpis:
        c1, c2 = cols[0], cols[-1]
        ws1.merge_cells(f"{c1}4:{c2}4")
        ws1.merge_cells(f"{c1}5:{c2}5")
        ws1.merge_cells(f"{c1}6:{c2}6")
        lc = ws1[f"{c1}4"]
        lc.value = label
        lc.font  = Font(bold=True, size=9, color=WHT, name="Calibri")
        lc.fill  = PatternFill("solid", fgColor=color)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        vc = ws1[f"{c1}5"]
        vc.value = val
        vc.font  = Font(bold=True, size=18, color=color, name="Calibri")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[5].height = 36
        ws1[f"{c1}6"].fill = PatternFill("solid", fgColor="EDEDED")

    # Cleaning log table
    ws1.merge_cells("A7:I7")
    ws1["A7"] = "  DATA CLEANING LOG"
    ws1["A7"].font = Font(bold=True, size=11, color=NAVY, name="Calibri")
    ws1["A7"].fill = PatternFill("solid", fgColor="D9E1F2")

    cl_hdr = ["Step", "Records Affected", "% of Pre-Step Rows", "Action"]
    for i, h in enumerate(cl_hdr, 1):
        c = ws1.cell(8, i, h)
        c.font = Font(bold=True, color=WHT, name="Calibri", size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _bdr
    ws1.row_dimensions[8].height = 24

    for i, (step, n, pct, action) in enumerate(clean_log, 9):
        bg = GRY if i % 2 == 0 else WHT
        for j, val in enumerate([step, n, pct, action], 1):
            cell = ws1.cell(i, j, val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.border    = _bdr
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left" if j in [1, 4] else "center", vertical="center")
            if isinstance(val, int) and j == 2:
                cell.number_format = "#,##0"

    # Anomaly breakdown table
    ar = 9 + len(clean_log) + 1
    ws1.merge_cells(f"A{ar}:I{ar}")
    ws1.cell(ar, 1, "  ANOMALY DETECTION BREAKDOWN")
    ws1[f"A{ar}"].font = Font(bold=True, size=11, color=NAVY, name="Calibri")
    ws1[f"A{ar}"].fill = PatternFill("solid", fgColor="D9E1F2")

    for i, h in enumerate(["Method", "Lower Fence / Threshold", "Upper Fence / Threshold",
                            "Flagged Records", "Rate (%)"], 1):
        c = ws1.cell(ar + 1, i, h)
        c.font      = Font(bold=True, color=WHT, name="Calibri", size=10)
        c.fill      = PatternFill("solid", fgColor=BLU)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _bdr
    ws1.row_dimensions[ar + 1].height = 24

    # composite row
    full_anom_log = anom_log + [("COMPOSITE (Union of All Methods)", "—", "—", total_anom, f"{anom_pct:.2f}%")]
    for i, (method, lo, hi, n, pct) in enumerate(full_anom_log, ar + 2):
        is_total = "COMPOSITE" in method
        bg = "FFF2CC" if is_total else (GRY if i % 2 == 0 else WHT)
        for j, val in enumerate([method, lo, hi, n, pct], 1):
            cell = ws1.cell(i, j, val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.border    = _bdr
            cell.font      = Font(bold=is_total, name="Calibri", size=10,
                                  color=RED if is_total else "000000")
            cell.alignment = Alignment(horizontal="left" if j == 1 else "center", vertical="center")
            if isinstance(val, int):
                cell.number_format = "#,##0"

    ws1.column_dimensions["A"].width = 32
    for col in "BCDEFGHI":
        ws1.column_dimensions[col].width = 18

    # ── Sheets 2–7 ────────────────────────────────────────────────────────────
    for name, df_s, color in [
        ("Monthly_KPI",      monthly,       GRN),
        ("Country_KPI",      country,       PUR),
        ("Top_50_Products",  top_products, "00B050"),
        ("Anomaly_By_Month", anom_month,   RED),
    ]:
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = color
        _write_df(ws, df_s, bg=color)
        _autofit(ws)
        ws.freeze_panes = "A2"

    # Color scale on Monthly Revenue
    n_m = len(monthly) + 1
    wb["Monthly_KPI"].conditional_formatting.add(
        f"B2:B{n_m}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max",   end_color=BLU),
    )
    # Color scale on Anomaly Rate
    n_a = len(anom_month) + 1
    wb["Anomaly_By_Month"].conditional_formatting.add(
        f"D2:D{n_a}",
        ColorScaleRule(start_type="min", start_color="E2EFDA",
                       end_type="max",   end_color="FF0000"),
    )

    # Anomaly Records sheet (color-coded by score)
    ws_ar = wb.create_sheet("Anomaly_Records")
    ws_ar.sheet_view.showGridLines = False
    ws_ar.sheet_properties.tabColor = RED
    _write_df(ws_ar, anom_exp, bg=RED)
    score_col_idx = anom_exp.columns.tolist().index("Anomaly_Score") + 1
    for r in range(2, len(anom_exp) + 2):
        score = ws_ar.cell(r, score_col_idx).value
        if   score and int(score) >= 5: bg = "FF6B6B"
        elif score and int(score) >= 3: bg = "FFCCCC"
        elif score and int(score) == 2: bg = "FFEEBA"
        else:                            bg = "FFFDE7"
        for c in range(1, len(anom_exp.columns) + 1):
            ws_ar.cell(r, c).fill = PatternFill("solid", fgColor=bg)
    _autofit(ws_ar, mx=28)
    ws_ar.freeze_panes = "A2"

    # Cleaned Data Sample
    ws_cd = wb.create_sheet("Cleaned_Data_Sample")
    ws_cd.sheet_view.showGridLines = False
    ws_cd.sheet_properties.tabColor = GRN
    sample_cols = ["InvoiceNo","StockCode","Description","Quantity","UnitPrice",
                   "Revenue","InvoiceDate","Country","CustomerID","InvoiceMonth","Quarter"]
    sdf = df_clean[sample_cols].head(2000).copy()
    sdf["InvoiceDate"] = sdf["InvoiceDate"].astype(str)
    _write_df(ws_cd, sdf, bg=GRN)
    _autofit(ws_cd)
    ws_cd.freeze_panes = "A2"

    wb.save(xlsx_path)
    logger.info(f"  [export] Excel Report  →  {xlsx_path}  ({len(wb.sheetnames)} sheets)")
    return xlsx_path


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    logger.info("=" * 60)
    logger.info("ERP DATA PIPELINE  –  START")
    logger.info("=" * 60)

    df_raw = load_data(INPUT_FILE)
    initial_rows = len(df_raw)

    df_clean, clean_log = clean_data(df_raw)
    logger.info(f"Rows after cleaning    →  {len(df_clean):,}")

    df_feat = engineer_features(df_clean)

    df_anom, anom_log, total_anom, anom_pct = detect_anomalies(df_feat)
    logger.info(f"Anomalies detected     →  {total_anom:,}  ({anom_pct:.2f} %)")

    xlsx = export_outputs(df_anom, clean_log, anom_log, total_anom, anom_pct, initial_rows)

    logger.info("=" * 60)
    logger.info("ERP DATA PIPELINE  –  COMPLETE ✓")
    logger.info(f"  Raw rows      : {initial_rows:,}")
    logger.info(f"  Cleaned rows  : {len(df_anom):,}")
    logger.info(f"  Anomalies     : {total_anom:,}  ({anom_pct:.2f}%)")
    logger.info(f"  Excel report  : {xlsx.name}")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
